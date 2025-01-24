import re
import click
import pandas as pd
import argparse
from os import listdir, makedirs
from os.path import isfile, join, exists
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from xlsx2html import xlsx2html
import imgkit
import tempfile
from datetime import date

# Argument parsing
parser = argparse.ArgumentParser(description='Health report parser')
parser.add_argument('--path', required=True, help='Path to health scripts folder')
args = parser.parse_args()

def text_file_to_series(file_path):
    data = {}
    with open(file_path, 'r') as file:
        for line in file:
            if "Details" in line:
                break
            if ': ' in line:
                all_fields = re.split(': |,', line)
                for field_name, value in zip(all_fields[::2], all_fields[1::2]):
                    data[field_name.strip()] = value.strip()
    return data

def clean_header(header_value):
    # Exclude "health-" and ".txt" from the header value
    header_value = header_value.replace("health-", "").replace(".txt", "")
    return header_value

@click.command()
@click.option('--path', default='.', help='Path where health scripts are present')
def build_summary(path):
    # Create a directory with today's date for storing files
    today = date.today().strftime("%Y-%m-%d")
    output_folder = f"health_summary_{today}"
    if not exists(output_folder):
        makedirs(output_folder)

    all_files = [f for f in listdir(path) if isfile(join(path, f))]
    all_health = {file: text_file_to_series(join(path, file)) for file in all_files}
    df = pd.DataFrame(all_health)
    
    # Create a workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Health Summary"
    
    # Define font style and fill style for the first row and column
    consolas_bold_font = Font(name='Consolas', bold=True)
    custom_fill = PatternFill(start_color='FF007BB8', end_color='FF007BB8', fill_type='solid')

    # Write DataFrame to the sheet starting from row 1
    for r_idx, r in enumerate(dataframe_to_rows(df, index=True, header=True), start=1):
        if r_idx == 2:  # Skip the empty row for the index
            continue
        for c_idx, value in enumerate(r, start=1):
            if r_idx == 1 and isinstance(value, str):
                value = clean_header(value)  # Clean up header value
            cell = ws.cell(row=r_idx - 1 if r_idx > 2 else r_idx, column=c_idx, value=value)
            # Apply font style
            cell.font = consolas_bold_font
            # Apply custom fill to the first row and first column
            if r_idx == 1 or c_idx == 1:
                cell.fill = custom_fill
            # Adding border to each cell
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
            cell.border = thin_border

    # Adjust column widths to fit data
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Function to extract numeric value from a cell
    def extract_numeric_value(cell_value):
        numeric_str = re.findall(r"[-+]?\d*\.\d+|\d+", str(cell_value))
        return float(numeric_str[0]) if numeric_str else None

    # Define fill styles for conditional formatting
    custom_green_fill = PatternFill(start_color='FF6AA84F', end_color='FF6AA84F', fill_type='solid')
    custom_red_fill = PatternFill(start_color='FFE06666', end_color='FFE06666', fill_type='solid')

    # Define keyword-based conditional formatting rules
    keyword_formats = {
        'Runs': {'threshold': 5, 'fill': custom_red_fill},
        'Missing': {'threshold': 1000, 'fill': custom_red_fill},
        'Reboots': {'threshold': 3, 'fill': custom_red_fill},
        'Runs ending in error': {'threshold': 2, 'fill': custom_red_fill},
        'USB reset events': {'threshold': 3, 'fill': custom_red_fill},
        'Distance': {'threshold': 5000, 'fill': custom_green_fill},
        'Max CPU temp': {'threshold': 82.0, 'fill': custom_red_fill},
        'Max WiFi temp': {'threshold': 90.0, 'fill': custom_red_fill},
        'Frames': {'threshold': 10000, 'fill': custom_green_fill},
        'Max Memory': {'threshold': 10.000, 'fill': custom_red_fill},
    }

    # Apply conditional formatting based on the first column values
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        keyword = row[0].value
        if keyword in keyword_formats:
            format_rule = keyword_formats[keyword]
            threshold = format_rule['threshold']
            fill = format_rule['fill']
            for cell in row[1:]:
                numeric_value = extract_numeric_value(cell.value)
                if numeric_value is not None and numeric_value > threshold:
                    cell.fill = fill

    # Save the workbook
    output_file = join(output_folder, "health_summary.xlsx")
    wb.save(output_file)
    print(f"Excel file '{output_file}' successfully created.")
    
    # Convert Excel to HTML
    html_file = join(output_folder, "health_summary.html")
    xlsx2html(output_file, html_file)
    print(f"HTML file '{html_file}' successfully created.")
    
    # Add Consolas font and custom fill to the HTML content
    with open(html_file, 'r', encoding='utf-8') as file:
        html_content = file.read()

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{
                font-family: Consolas, monospace;
                font-weight: bold;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                border: 1px solid black;
                padding: 8px;
                text-align: left;
                font-family: Consolas, monospace;
                font-weight: bold;
            }}
            th {{
                background-color: #007BB8;
                color: #FFFFFF;
            }}
            tr:first-child td, tr td:first-child {{
                background-color: #007BB8;
                color: #FFFFFF;
            }}
        </style>
    </head>
    <body>
        {html_content}
    </body>
    </html>
    """

    with open(html_file, 'w', encoding='utf-8') as file:
        file.write(html_content)
    
    # Convert HTML to JPEG
    jpeg_file = join(output_folder, "health_summary.jpeg")
    
    # Create a temporary HTML file to get its height
    temp_html = tempfile.NamedTemporaryFile(delete=False, suffix='.html')
    temp_html.write(html_content.encode('utf-8'))
    temp_html.close()

    # Get height of the HTML content
    options = {
        'format': 'jpeg',
        'quality': '100',
        'width': 1500,
        'disable-smart-width': '',
        'custom-header': [
            ('Accept-Encoding', 'gzip')
        ]
    }
    imgkit.from_file(temp_html.name, jpeg_file, options=options)
    print(f"JPEG file '{jpeg_file}' successfully created.")

    # Clean up temporary HTML file
    import os
    os.remove(temp_html.name)

if __name__ == "__main__":
    build_summary()

